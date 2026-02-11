"""
Main orchestration pipeline for disclosure change analysis.
"""
import argparse
import json
import logging
from pathlib import Path
import pandas as pd
from tqdm import tqdm

from .parser import parse_all_pdfs, save_parsed_data
from .analyzer import analyze_all_companies

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


def save_styled_excel(df: pd.DataFrame, output_path: Path):
    """
    Save DataFrame to an Excel file with colored columns and headers.
    """
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Disclosure Changes"
    
    # Define column color mapping (ARGB hex codes)
    # Using light pastel colors for readability
    column_colors = {
        "Company": "E3F2FD",           # Light Blue
        "Quarter_Previous": "F3E5F5",  # Light Purple
        "Quarter_Current": "F3E5F5",   # Light Purple
        "Section": "E8F5E9",           # Light Green
        "Quote_Old": "FFF3E0",         # Light Orange
        "Quote_New": "FFF3E0",         # Light Orange
        "Description": "F1F8E9",      # Light Yellow/Green
        "Signal": "FFEBEE"             # Light Red/Pink
    }
    
    thin_border = Border(
        left=Side(style='thin', color="CCCCCC"),
        right=Side(style='thin', color="CCCCCC"),
        top=Side(style='thin', color="CCCCCC"),
        bottom=Side(style='thin', color="CCCCCC")
    )
    
    headers = list(df.columns)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        color = column_colors.get(header, "FFFFFF")
        
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        
        if "Quote" in header or "Description" in header:
            ws.column_dimensions[cell.column_letter].width = 60
        else:
            ws.column_dimensions[cell.column_letter].width = 18
            
    for row_num, row_data in enumerate(df.values, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=str(value))
            header = headers[col_num - 1]
            color = column_colors.get(header, "FFFFFF")
            
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = thin_border
            
    ws.freeze_panes = "A2"
    wb.save(output_path)
    logger.info(f"Saved styled Excel report to {output_path}")

def run_pipeline(
    data_dir: str = "data",
    output_dir: str = "output",
    dry_run: bool = False,
    skip_parsing: bool = False,
    use_semantic: bool = True,
    target_company: str = None,
    skip_file_output: bool = False,
    run_evaluation: bool = False
):
    """
    Run the complete disclosure analysis pipeline.
    
    Args:
        data_dir: Directory containing PDF files
        output_dir: Directory for output files
        dry_run: If True, skip LLM API calls (for testing)
        skip_parsing: If True, use existing parsed_data.json
        use_semantic: If True, use AI-based semantic extraction (for earnings transcripts).
                     If False, use regex-based extraction (for structured SEC filings).
        target_company: Optional filter for specific company files.
        skip_file_output: If True, skip CSV/Excel generation (for UI performance)
        run_evaluation: If True, run judge LLM evaluation after analysis
    """
    import time
    start_time = time.time()
    timings = {}
    summary = {"results": [], "verdict": None, "usage": {}}

    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)
    
    parsed_data_path = output_path / "parsed_data.json"
    
    # Step 1: Parse PDFs
    step1_start = time.time()
    
    # Always try to load existing data first to enable incremental updates
    existing_data = {}
    if parsed_data_path.exists():
        try:
            with open(parsed_data_path, 'r') as f:
                existing_data = json.load(f)
            logger.info(f"Loaded existing data for {len(existing_data)} companies")
        except Exception as e:
            logger.error(f"Failed to load existing parsed data: {e}")

    if skip_parsing:
        parsed_data = existing_data
    else:
        logger.info(f"\n{'='*60}")
        logger.info("STEP 1: Parsing PDF Documents")
        if target_company:
            logger.info(f"Targeting company: {target_company}")
        logger.info(f"{'='*60}\n")
        
        new_parsed_data = parse_all_pdfs(
            data_dir, 
            use_semantic_extraction=use_semantic,
            target_company=target_company,
            existing_data=existing_data
        )
        
        # Merge logic: Update existing data with new data
        # If target_company is set, we only update that company's data
        parsed_data = existing_data.copy()
        for company, quarters in new_parsed_data.items():
            if company not in parsed_data:
                parsed_data[company] = {}
            # Update specific quarters (overwrite if exists, add if new)
            for q, text in quarters.items():
                parsed_data[company][q] = text
        
        if not parsed_data:
            logger.error("No data parsed (and no existing data). Check your PDF files.")
            # return {"results": [], "verdict": None, "usage": {}} # Don't return yet, maybe cache has something?
        
        save_parsed_data(parsed_data, str(parsed_data_path))
        
        # Print summary
        total_quarters = sum(len(quarters) for quarters in parsed_data.values())
        logger.info(f"\nParsing Summary:")
        logger.info(f"  Extraction mode: {'Semantic (AI)' if use_semantic else 'Regex (SEC filings)'}")
        logger.info(f"  Companies: {len(parsed_data)}")
        logger.info(f"  Total quarters: {total_quarters}")
        for company, quarters in parsed_data.items():
            logger.info(f"  {company}: {sorted(quarters.keys())}")
    
    timings['Step 1: Parsing & Extraction'] = time.time() - step1_start

    # Step 2: Analyze with LLM
    step2_start = time.time()
    logger.info(f"\n{'='*60}")
    logger.info("STEP 2: LLM-based Disclosure Analysis")
    logger.info(f"{'='*60}\n")
    
    if dry_run:
        logger.warning("DRY RUN MODE: Will not make actual LLM API calls")
    
    analysis_data, usage = analyze_all_companies(parsed_data, dry_run=dry_run)
    changes = analysis_data.get("results", [])
    verdict_data = analysis_data.get("verdict")
    
    timings['Step 2: LLM Analysis'] = time.time() - step2_start

    if not changes and not dry_run:
        logger.warning("No changes detected across all companies")
        return {"results": [], "verdict": None, "usage": usage}
    
    # Step 3: Generate output
    step3_start = time.time()
    logger.info(f"\n{'='*60}")
    logger.info("STEP 3: Generating Output")
    logger.info(f"{'='*60}\n")
    
    if not dry_run:
        df = pd.DataFrame(changes)

        if not skip_file_output:
            # Save to CSV
            csv_path = output_path / "disclosure_changes.csv"
            df.to_csv(csv_path, index=False)
            logger.info(f"Saved {len(df)} changes to {csv_path}")

            # Save to Styled Excel
            excel_path = output_path / "disclosure_changes.xlsx"
            try:
                save_styled_excel(df, excel_path)
            except Exception as e:
                logger.error(f"Failed to save styled Excel: {e}")

        # Display Final Verdict if available
        if verdict_data:
            logger.info("\n" + "="*60)
            logger.info("FINAL ANALYSIS VERDICT")
            logger.info("="*60)
            logger.info(f"SIGNAL: {verdict_data.get('final_signal', 'N/A').upper()}")
            logger.info("-" * 60)
            logger.info("INSIGHTS & HIGHLIGHTS:")
            logger.info(verdict_data.get('insights', 'No insights generated.'))
            logger.info("-" * 60)
            logger.info("STRATEGIC VERDICT:")
            logger.info(verdict_data.get('verdict', 'No verdict generated.'))
            logger.info("=" * 60 + "\n")

        # Generate summary statistics
        logger.info("\n" + "="*60)
        logger.info("SUMMARY STATISTICS")
        logger.info("="*60)
        logger.info(f"Total changes detected: {len(df)}")
        logger.info(f"\nBy signal classification:")
        for signal, count in df['Signal'].value_counts().items():
            logger.info(f"  {signal}: {count}")
        logger.info(f"\nBy section:")
        for section, count in df['Section'].value_counts().items():
            logger.info(f"  {section}: {count}")
        
        # Build summary
        summary = {
            "total_changes": len(df),
            "by_signal": df['Signal'].value_counts().to_dict(),
            "by_section": df['Section'].value_counts().to_dict(),
            "verdict": verdict_data,
            "usage": usage
        }
        if not skip_file_output:
            summary_path = output_path / "summary.json"
            with open(summary_path, 'w') as f:
                json.dump(summary, f, indent=2)
            logger.info(f"Saved summary to {summary_path}")
    
    timings['Step 3: Output Generation'] = time.time() - step3_start
    total_time = time.time() - start_time

    # Cost Calculation (Gemini 1.5 Flash Prices)
    # $0.10 / 1M input tokens, $0.40 / 1M output tokens
    input_tokens = usage.get("input_tokens", 0)
    output_tokens = usage.get("output_tokens", 0)
    estimated_cost = (input_tokens * 0.10 + output_tokens * 0.40) / 1_000_000

    logger.info("\n" + "="*60)
    logger.info("AI COST & PERFORMANCE LOGS")
    logger.info("="*60)
    logger.info(f"{'Input Tokens':.<40} {input_tokens:,}")
    logger.info(f"{'Output Tokens':.<40} {output_tokens:,}")
    logger.info(f"{'Estimated Cost (USD)':.<40} ${estimated_cost:.4f}")
    logger.info("-" * 50)
    for step, duration in timings.items():
        logger.info(f"{step:.<40} {duration:>8.2f}s")
    logger.info("-" * 50)
    logger.info(f"{'TOTAL PIPELINE TIME':.<40} {total_time:>8.2f}s")
    logger.info("=" * 60)

    # Step 4: Judge LLM Evaluation (Optional)
    if run_evaluation and not dry_run and changes:
        step4_start = time.time()
        logger.info(f"\n{'='*60}")
        logger.info("STEP 4: Judge LLM Evaluation")
        logger.info(f"{'='*60}\n")
        
        try:
            from .judge_evaluator import run_full_evaluation
            from .evaluation_report import generate_evaluation_report
            
            evaluation_results = run_full_evaluation(
                parsed_data=parsed_data,
                analysis_results=changes,
                verdict=verdict_data
            )
            
            # Generate reports
            generate_evaluation_report(evaluation_results, output_dir)
            
            # Add to summary
            summary["evaluation"] = evaluation_results.get("summary_scores", {})
            
            timings['Step 4: Judge Evaluation'] = time.time() - step4_start
        except Exception as e:
            logger.error(f"Evaluation failed: {e}")
            logger.exception(e)

    logger.info("\n" + "="*60)
    logger.info("PIPELINE COMPLETE!")
    logger.info("="*60 + "\n")

    return summary



def main():
    """CLI entry point."""
    parser = argparse.ArgumentParser(
        description="Financial Disclosure Change Analysis Pipeline"
    )
    parser.add_argument(
        "--data-dir",
        default="data",
        help="Directory containing PDF files (default: data)"
    )
    parser.add_argument(
        "--output-dir",
        default="output",
        help="Directory for output files (default: output)"
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Skip LLM API calls (for testing)"
    )
    parser.add_argument(
        "--skip-parsing",
        action="store_true",
        help="Use existing parsed_data.json if available"
    )
    parser.add_argument(
        "--use-regex",
        action="store_true",
        help="Use regex-based extraction (for structured SEC filings). Default is semantic extraction (for earnings transcripts)."
    )
    parser.add_argument(
        "--evaluate",
        action="store_true",
        help="Run judge LLM evaluation after analysis to assess accuracy"
    )
    
    args = parser.parse_args()
    
    run_pipeline(
        data_dir=args.data_dir,
        output_dir=args.output_dir,
        dry_run=args.dry_run,
        skip_parsing=args.skip_parsing,
        use_semantic=not args.use_regex,  # Default to semantic
        run_evaluation=args.evaluate
    )


if __name__ == "__main__":
    main()
